import argparse
import subprocess
import os
import sys
import itertools
import time
from threading import Thread, Event
from colorama import Fore, Style
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# -----------------------------------------------------------
# This script automates scanning a container image using Syft,
# extracts component information, and generates an Excel file
#
# This script was created by Aldo Leon with help from ChatGPT.
# -----------------------------------------------------------

# Ensure the Output directory exists
output_dir = "Output"
os.makedirs(output_dir, exist_ok=True)

# Animation for processing
def animate_processing(message, stop_event):
    animation = itertools.cycle(['|', '/', '~', '\\'])
    print(f"{Fore.YELLOW}{message}{Style.RESET_ALL}", end="", flush=True)
    while not stop_event.is_set():
        sys.stdout.write(next(animation))
        sys.stdout.flush()
        time.sleep(0.5)
        sys.stdout.write("\b")

# Function to run Syft and capture its output
def run_syft(image_id):
    try:
        command = ["syft", image_id, "-o", "json"]
        print(f"{Fore.YELLOW}Running Syft for image: {image_id}{Style.RESET_ALL}")

        # Set up animation thread
        stop_event = Event()
        animation_thread = Thread(target=animate_processing, args=("Processing image with Syft: ", stop_event))
        animation_thread.start()

        # Run Syft command
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stop_event.set()  # Stop animation
        animation_thread.join()

        if result.returncode != 0:
            print(f"{Fore.RED}Error running Syft: {result.stderr}{Style.RESET_ALL}")
            sys.exit(1)
        return result.stdout
    except Exception as e:
        print(f"{Fore.RED}An error occurred: {e}{Style.RESET_ALL}")
        sys.exit(1)

# Function to write Excel (XLSX) from Syft JSON output (without Licenses column)
def write_excel(image_name,image_id, syft_output):
    output_filename = os.path.join(output_dir, f"Components_found_in_{image_name}.xlsx")
    components = []

    try:
        import json
        data = json.loads(syft_output)
        for artifact in data.get("artifacts", []):
            components.append({
                "Name": artifact.get("name"),
                "Version": artifact.get("version"),
                "Type": artifact.get("type"),
            })
    except json.JSONDecodeError as e:
        print(f"{Fore.RED}Failed to parse Syft JSON output: {e}{Style.RESET_ALL}")
        sys.exit(1)

    # Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Components"

    # Define color formatting
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow strong
    data_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")    # Light yellow
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Write headers with formatting
    headers = ["Name", "Version", "Type"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border

    # Write data rows with formatting
    for row_num, component in enumerate(components, start=2):
        for col_num, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=component[key])
            cell.fill = data_fill
            cell.border = border

    # Add a formatted summary of findings from E3 to H7
    summary_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold for summary
    bold_font_summary = Font(bold=True)
    center_alignment_summary = Alignment(horizontal="center", vertical="center")

    summary_data = [
        ("Total Components Found:", len(components)),
    ]

    # Writing the summary
    ws['E3'] = "Summary of Findings"
    ws['E3'].font = bold_font_summary
    ws['E3'].fill = summary_fill
    ws['E3'].alignment = center_alignment_summary

    row_num = 4
    for label, value in summary_data:
        ws[f'E{row_num}'] = label
        ws[f'F{row_num}'] = value
        ws[f'E{row_num}'].font = bold_font_summary
        ws[f'F{row_num}'].font = bold_font_summary
        ws[f'E{row_num}'].fill = summary_fill
        ws[f'F{row_num}'].fill = summary_fill
        ws[f'E{row_num}'].alignment = center_alignment_summary
        ws[f'F{row_num}'].alignment = center_alignment_summary
        ws[f'E{row_num}'].border = border
        ws[f'F{row_num}'].border = border
        row_num += 1

    # Adjust column width based on content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get column name (e.g. 'A', 'B', etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add some padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the Excel file
    wb.save(output_filename)

    print(f"{Fore.GREEN}Excel file created: {output_filename}{Style.RESET_ALL}")

    # Provide a summary of the findings
    print(f"\n{Fore.CYAN}Summary of Findings:{Style.RESET_ALL}")
    print(f"- Number of components found: {len(components)}")
    print(f"- Path of the Excel file: {os.path.abspath(output_filename)}")

# Main function
def main():
    parser = argparse.ArgumentParser(description="Run Syft and output components to Excel.")
    parser.add_argument("-id", "--image-id", required=True, help="The image identifier to scan.")
    parser.add_argument("-name", "--image-name", required=True, help="The image name to scan.")
    args = parser.parse_args()

    image_id = args.image_id
    image_name = args.image_name

    print(f"{Fore.YELLOW}Starting Syft scan for image: {image_name}{Style.RESET_ALL}")

    # Step 1: Run Syft
    print(f"{Fore.YELLOW}Step 1: Running Syft...{Style.RESET_ALL}")
    syft_output = run_syft(image_id)

    # Step 2: Write to Excel
    print(f"{Fore.YELLOW}Step 2: Writing results to Excel...{Style.RESET_ALL}")
    write_excel(image_name, image_id, syft_output)

    print(f"{Fore.GREEN}Process completed successfully!{Style.RESET_ALL}")

if __name__ == "__main__":
    main()

